"""Формирование отчёта о найденных вхождениях."""
from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import List

import pandas as pd

from notice import full_notice
from utils.logger import get_logger

log = get_logger()

# Для построения «гибкого» поиска контекста в исходном тексте.
_WORD = r"[0-9A-Za-zА-Яа-яЁё]"


@dataclass
class Finding:
    """Одно совпадение для отчёта."""

    entity: str        # имя/организация как в перечне
    category: str      # категория из перечня
    confidence: str    # достоверность (высокая/средняя/низкая)
    page_url: str      # где найдено
    page_title: str    # заголовок страницы
    context: str       # фрагмент текста вокруг совпадения


def extract_context(original_text: str, norm_pattern: str, width: int = 80) -> str:
    """Найти фрагмент исходного текста вокруг совпадения.

    Строит гибкий regex из токенов нормализованного имени (любые разделители
    между словами, без учёта регистра) и возвращает контекст ±``width``.
    """
    tokens = [re.escape(t) for t in norm_pattern.split(" ") if t]
    if not tokens:
        return ""
    flexible = r"[^0-9A-Za-zА-Яа-яЁё]+".join(tokens)
    pattern = rf"(?<!{_WORD})(?:{flexible})(?!{_WORD})"
    m = re.search(pattern, original_text, re.IGNORECASE)
    if not m:
        return ""
    start = max(0, m.start() - width)
    end = min(len(original_text), m.end() + width)
    snippet = re.sub(r"\s+", " ", original_text[start:end]).strip()
    return f"…{snippet}…"


def write_report(
    findings: List[Finding], directory: str, basename: str
) -> List[Path]:
    """Сохранить отчёт в CSV и XLSX. Возвращает пути созданных файлов."""
    out_dir = Path(directory)
    out_dir.mkdir(parents=True, exist_ok=True)

    columns = {
        "entity": "Имя/организация (перечень)",
        "category": "Категория",
        "confidence": "Достоверность",
        "page_url": "Страница (URL)",
        "page_title": "Заголовок страницы",
        "context": "Контекст",
    }
    # Сортируем: сначала высокодостоверные совпадения.
    order = {"высокая": 0, "средняя": 1}
    findings = sorted(findings, key=lambda f: order.get(f.confidence, 2))

    df = pd.DataFrame([asdict(f) for f in findings])
    if df.empty:
        df = pd.DataFrame(columns=list(columns))
    else:
        df = df.rename(columns=columns)[list(columns.values())]

    created: List[Path] = []
    csv_path = out_dir / f"{basename}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    created.append(csv_path)

    xlsx_path = out_dir / f"{basename}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="matches")
        _style_worksheet(writer.sheets["matches"], df)
        # Отдельный лист с правовой оговоркой и происхождением данных.
        info = pd.DataFrame({"Правовая оговорка / источники": full_notice().splitlines()})
        info.to_excel(writer, index=False, sheet_name="Инфо")
        writer.sheets["Инфо"].column_dimensions["A"].width = 100
    created.append(xlsx_path)

    log.info("Отчёт сохранён: %s (совпадений: %d)", basename, len(findings))
    return created


# Заливка ячеек «Достоверности» по уровню (как «светофор» Excel).
_CONF_FILL = {
    "высокая": "C6EFCE",   # зелёный
    "средняя": "FFEB9C",   # жёлтый
    "низкая": "FFC7CE",    # красный
}
# Разумные пределы ширины колонок (символы).
_MAX_WIDTH = {"Контекст": 70, "Заголовок страницы": 35, "Страница (URL)": 50}


def _style_worksheet(ws, df: pd.DataFrame) -> None:
    """Оформить лист: жирная шапка, автоширина, перенос, цвет достоверности."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Шапка: жирная, с заливкой, закрепляем первую строку.
    for col_idx, _ in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    conf_col = (
        list(df.columns).index("Достоверность") + 1
        if "Достоверность" in df.columns
        else None
    )

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)

        # Автоподбор ширины по самой длинной строке (с разумным пределом).
        values = [str(col_name)] + [str(v) for v in df[col_name].tolist()]
        longest = max((len(v) for v in values), default=10)
        width = min(longest + 2, _MAX_WIDTH.get(col_name, 45))
        ws.column_dimensions[letter].width = width

        # Перенос текста в длинных колонках.
        if col_name in _MAX_WIDTH:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = wrap

    # Заливка ячеек колонки «Достоверность» по уровню уверенности.
    if conf_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=conf_col)
            level = str(cell.value).split(" ", 1)[0]  # «низкая (…)» -> «низкая»
            color = _CONF_FILL.get(level)
            if color:
                cell.fill = PatternFill("solid", fgColor=color)
